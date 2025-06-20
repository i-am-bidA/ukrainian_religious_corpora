import time
import re
import pandas as pd
from collections import Counter, defaultdict
from pathlib import Path
from pymorphy2 import MorphAnalyzer
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

morph = MorphAnalyzer(lang='uk')

with open("stop_words.txt", encoding="utf-8") as f:
    stop_words = set(line.strip().lower() for line in f if line.strip() and len(line.strip().split()) == 1)

pcu_files = [
    "Новини і проповіді ПЦУ.txt",
    "ПЦУ Труди КДА 2022.txt",
    "ПЦУ Труди КДА 2023.txt",
    "ПЦУ Труди КДА 2024.txt",
    "Церква і суспільство випуск 1.txt",
    "Церква і суспільство випуск 2.txt",
    "Церква і суспільство випуск 3.txt",
    "Церква і суспільство випуск 4.txt",
    "Церква і суспільство випуск 5.txt"
]

upc_files = [
    "Новини УПЦ (телеграм).txt",
    "Проповіді Онуфрія УПЦ (телеграм).txt",
    "УПЦ Труди КДА випуск 38 (2023).txt",
    "УПЦ Труди КДА випуск 39 (2023).txt",
    "УПЦ Труди КДА випуск 40 (2024).txt",
    "УПЦ Труди КДА випуск 41 (2024).txt"
]

abbreviation_expansion = {
    "ап.": "апостол", "апп.": "апостоли", "архієп.": "архієпископ", "архім.": "архимандрит", "об.": "об’явлення",
    "безср.": "безсрібник", "блгв.": "благовірний", "блж.": "блаженний", "вмц.": "великомучениця",
    "вмч.": "великомученик", "дияк.": "диякон", "єп.": "єпископ", "ігум.": "ігумен", "спов.": "сповідник",
    "кн.": "князь", "митр.": "митрополит", "мц.": "мучениця", "мцц.": "мучениць", "мч.": "мученик",
    "мчч.": "мучеників", "патр.": "патріарх", "прав.": "праведний", "прп.": "преподобний", "прпп.": "преподобні",
    "прпмц.": "преподобномучениця", "прпмч.": "преподобномученик", "пресвіт.": "пресвітер", "прор.": "пророк",
    "св.": "святий", "свв.": "святі", "свт.": "святитель", "свтт.": "святителі", "сщмч.": "священномученик",
    "сщмчн.": "священномучениця", "сщч.": "священнослужитель", "вч.": "вчитель", "ввеч.": "вечірня", "с.": "сторінка",
    "ран.": "рання", "літ.": "літургія", "мф.": "від матфея", "мк.": "від марка", "лк.": "від луки", "іуд.": "іуда",
    "ін.": "від іоана / інше", "діян.": "діяння святих апостолів", "як.": "послання апостола якова", "пет.": "послання апостола петра",
    "кор.": "до коринфян", "рим.": "до римлян", "гал.": "до галатів", "еф.": "до ефесян", "флп.": "до филип’ян", "кол.": "до колосян", "прот.": "протоієрей",
    "сол.": "до солунян", "тим.": "до тимофія", "архімм.": "архімандрити", "блгвв.": "благовірні", "блжж.": "блаженні",
    "тит.": "до тита", "флм.": "до филимона", "євр.": "до євреїв", "свящ.": "священник", "притч.": "притчі соломона",
    "бут.": "буття", "вих.": "вихід", "іс.": "ісая", "іоїл.": "йоїль", "зах.": "захарія", "архієпп.": "архієпископи",     
    "вел.": "великий, велика", "вмцц.": "великомучениці", "вмчч.": "великомученики", "єв.": "євангеліст",
    "єпп.": "єпископи", "ієром.": "ієромонах", "ієросхим.": "ієросхимонах", "імп.": "імператор",
    "кнн.": "князі", "кнг.": "княгиня", "кнж.": "княжна", "митрр.": "митрополити", "новмч.": "новомученик",
    "новосвщмч.": "новосвященномученик", "патрр.": "патріархи", "правв.": "праведні", "прорр.": "пророки",
    "пророчц.": "пророчиця", "просвіт.": "просвітитель, просвітителька", "протопресв.": "протопресвітер",
    "прмч.": "преподобномученик", "прмчч.": "преподобномученики", "прмц.": "преподобномучениця", "прмцц.": "преподобномучениці",
    "рівноап.": "рівноапостольний, рівноапостольна", "рівноапп.": "рівноапостольні", "сп.": "сповідник, сповідниця",
    "сщмчч.": "священномученики", "стовпн.": "стовпник", "страст.": "страстотерпець", "схим.": "схимонах",
    "чудотв.": "чудотворець", "юрод.": "юродивий", "бр.": "братчик", "вол.": "волость", "впцр": "Всеукраїнська Православна Церковна Рада",
    "впцс": "Всеукраїнський Православний Церковний Собор", "вспп": "Всеукраїнська Спілка православних парафій",
    "вуцвк": "Всеукраїнський Центральний Виконавчий Комітет", "вуцик": "Всеукраїнський Центральний Виконавчий Комітет",
    "вцр": "Всеукраїнська Церковна Рада", "вцс": "Всеукраїнський Церковний Собор", "головноупов.": "головноуповноважений",
    "гр.": "громадянин", "д.": "добродій / доктор", "див.": "дивись", "док.": "документ", "кпцр": "Київська Православна Церковна Рада",
    "кгчк": "Київська губернська надзвичайна комісія", "мед.": "медичний", "ст.": "стиль", "о.": "отець", "окр.": "округ", "оп.": "опис",
    "п.": "пан", "пар.": "парафія", "параф.": "парафіяльний", "п.-о.": "пан-отець", "пов.": "повіт", 
    "пред.": "предстоятель", "проф.": "професор", "рпц": "Російська Православна Церква",
    "р. х.": "Різдво Христове", "с.-д.": "соціал-демократ", "совнарком": "Рада Народних Комісарів", "спр.": "справа",
    "ст.": "сторінка", "т-во": "товариство", "т. з.": "так званий", "т. зр.": "точка зору",
    "тов.": "товариш", "уапц": "Українська Автокефальна Православна Церква", "угкц": "Українська Греко-Католицька Церква",
    "укр.": "український", "унр": "Українська Народна Республіка", "упц": "Українська Православна Церква",
    "упцр": "Українська Православна Церковна Рада", "уск": "Українська Споживча Кооперація",
    "усрр": "Українська Соціалістична Радянська Республіка", "уцр": "Українська Центральна Рада", "ф.": "--",
    "хр.": "християнський", "цит.": "цитата", "ціж": "Церква і Життя", "цр": "Церковна Рада", "ч.": "число",
    "част.": "частина", "см.": "смотри / дивись", "м.": "місто", "н.": "наук", "філософ.": "філософський",
    "проф.": "професор", "ред.": "редактор / редакція", "богосл.": "богословський", "італ.": "італійською", "іс.": "ісая",
    "ім.": "імені", "вип.": "випуск", "арам.": "арамейський", "англ.": "англійською", "нав.": "Навин (книга Ісуса Навина)",
    "суд.": "книга Суддів Ізраїлевих", "чис.": "книга Числа", "укр.": "українською", "гр.": "грецький", "івр.": "іврит",
    "єз.": "пророк Єзекіїль", "філолог.": "філологічний", "арк.": "аркуш", "т.": "так", "зв.": "званий / зворот", 
}

def process_corpus(files):
        roman_re = re.compile(r"^(?=[MDCLXVI])M{0,4}(CM|CD|D?C{0,3})(XC|XL|L?X{0,3})(IX|IV|V?I{0,3})$", re.IGNORECASE)

        global_counter = Counter()
        source_counters = defaultdict(Counter)
        file_tracker = defaultdict(set)
        abbreviations = defaultdict(int)
        latin_words = defaultdict(int)
        all_tokens = []

        for file in files:
            path = Path(file)
            if not path.exists():
                continue

            with open(path, "r", encoding="utf-8") as f:
                text = f.read().lower()

            text = text.replace("’", "'").replace("ʼ", "'").replace("\u00AD", "-").replace("–", "-").replace("—", "-")
            text = re.sub(r"(\w+)-\s*\n\s*(\w+)", r"\1\2", text)

            # for word in re.findall(r"\b[a-zA-Z]{2,}\b", text):
            #     if not roman_re.fullmatch(word):
            #         latin_words[word] += 1
            #         file_tracker[f"LATIN::{word}"].add(file)

            abbrs = re.findall(r"\b[а-щґєіїьюя]{1,4}\.", text)
            for abbr in abbrs:
                if abbr in abbreviation_expansion:
                    abbreviations[abbr] += 1
                    file_tracker[f"ABBR::{abbr}"].add(file)

            raw_tokens = re.findall(r"\b[а-щґєіїьюя]+(?:[-'][а-щґєіїьюя]+)*\.?\b", text)

            i = 0
            tokens = []
            while i < len(raw_tokens):
                w1 = raw_tokens[i]
                if i + 1 < len(raw_tokens):
                    w2 = raw_tokens[i + 1]
                    combined = w1 + w2
                    if (not morph.word_is_known(w1) or not morph.word_is_known(w2)) and morph.word_is_known(combined):
                        tokens.append(combined)
                        file_tracker[combined].update(file_tracker[w1])
                        file_tracker[combined].update(file_tracker[w2])
                        i += 2
                        continue
                tokens.append(w1)
                i += 1

            for t in tokens:
                file_tracker[t].add(file)
            source_counters[file].update(tokens)
            global_counter.update(tokens)
            all_tokens.extend(tokens)

        main_rows = []
        lemma_pos_dict = {}
        lemma_counter = Counter()

        for word, freq in global_counter.items():
            parse = morph.parse(word)[0]
            lemma = parse.normal_form
            pos = parse.tag.POS or "-"
            file_list = ", ".join(file_tracker[word])
            main_rows.append((word, lemma, pos, freq, file_list))
            lemma_pos_dict[word] = (lemma, pos)
            lemma_counter[lemma] += freq

        main_df = pd.DataFrame(main_rows, columns=["Слово", "Лема", "Частина мови", "Частота", "Файли"])

        lemma_df = pd.DataFrame([
            (lemma, freq) for lemma, freq in lemma_counter.items()
        ], columns=["Лема", "Частота"])
        lemma_df["Відносна частота"] = lemma_df["Частота"] / sum(lemma_df["Частота"])

        top20_df = pd.DataFrame([
            (file, word, freq) for file, counter in source_counters.items()
            for word, freq in counter.most_common(25)
        ], columns=["Файл", "Слово", "Частота"])

        top20_content_df = pd.DataFrame([
            (file, word, freq) for file, counter in source_counters.items()
            for word, freq in sorted([(w, c) for w, c in counter.items()
            if w not in stop_words and lemma_pos_dict.get(w, ("", ""))[1] in {"NOUN", "VERB", "ADJ", "ADV", "PROPN"}],
            key=lambda x: -x[1])[:25]
        ], columns=["Файл", "Слово", "Частота"])

        latin_df = pd.DataFrame([
            (word, count, ", ".join(file_tracker[f"LATIN::{word}"])) for word, count in latin_words.items()
        ], columns=["Слово (латиниця)", "Частота", "Файли"])

        abbrev_df = pd.DataFrame([
            (abbr, count, abbreviation_expansion.get(abbr, "-"), ", ".join(file_tracker[f"ABBR::{abbr}"]))
            for abbr, count in abbreviations.items()
        ], columns=["Скорочення", "Частота", "Розшифрування", "Файли"])

        bigram_counter = Counter()
        trigram_counter = Counter()
        bigram_sources = defaultdict(set)
        trigram_sources = defaultdict(set)

        for i in range(len(all_tokens) - 1):
            bg = all_tokens[i] + " " + all_tokens[i + 1]
            bigram_counter[bg] += 1
            bigram_sources[bg].update(file_tracker[all_tokens[i]])

        for i in range(len(all_tokens) - 2):
            tg = all_tokens[i] + " " + all_tokens[i + 1] + " " + all_tokens[i + 2]
            trigram_counter[tg] += 1
            trigram_sources[tg].update(file_tracker[all_tokens[i]])

        bigrams_df = pd.DataFrame([
            (k, v, ", ".join(bigram_sources[k]), v / sum(bigram_counter.values()))
            for k, v in bigram_counter.items() if v >= 3
        ], columns=["2-грам", "Частота", "Файли", "Відносна частота"])

        trigrams_df = pd.DataFrame([
            (k, v, ", ".join(trigram_sources[k]), v / sum(trigram_counter.values()))
            for k, v in trigram_counter.items() if v >= 3
        ], columns=["3-грам", "Частота", "Файли", "Відносна частота"])

        return {
            "main_df": main_df,
            "lemma_df": lemma_df,
            "top20_df": top20_df,
            "top20_content_df": top20_content_df,
            "latin_df": latin_df,
            "abbrev_df": abbrev_df,
            "bigrams_df": bigrams_df,
            "trigrams_df": trigrams_df,
            "lemmas_set": set(lemma_counter.keys()),
            "bigrams_set": set(bigram_counter.keys()),
            "trigrams_set": set(trigram_counter.keys()),
        }

def autosize_excel_columns(writer, sheet_names):
    for sheet_name in sheet_names:
        worksheet = writer.sheets[sheet_name]
        for column_cells in worksheet.columns:
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = max_length + 2

# Функція формування таблиці порівняння
def make_common_df(units, col_name, r1, r2):
    df = pd.DataFrame(units, columns=[col_name])
    df["Частота (ПЦУ)"] = df[col_name].map(r1)
    df["Частота (УПЦ)"] = df[col_name].map(r2)
    df["Відносна (ПЦУ)"] = df["Частота (ПЦУ)"] / df["Частота (ПЦУ)"].sum()
    df["Відносна (УПЦ)"] = df["Частота (УПЦ)"] / df["Частота (УПЦ)"].sum()
    df.insert(0, "Номер", range(1, len(df)+1))
    return df

# Функціональність для обробки двох корпусів
def compare_and_save_corpora(files1, files2, filename1, filename2, output_file):
    result1 = process_corpus(files1)
    print("Обробив тексти ПЦУ...")
    result2 = process_corpus(files2)
    print("Обробив тексти УПЦ...\n")
    print('Переходжу до другого етапу')

    common_lemmas = sorted(result1["lemmas_set"] & result2["lemmas_set"])
    common_bigrams = sorted(result1["bigrams_set"] & result2["bigrams_set"])
    common_trigrams = sorted(result1["trigrams_set"] & result2["trigrams_set"])

    diff_lemmas = sorted(result1["lemmas_set"] ^ result2["lemmas_set"])

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        for label, res in [(filename1, result1), (filename2, result2)]:
            res["main_df"].sort_values(by="Частота", ascending=False).to_excel(writer, sheet_name=f"{label}_Слова", index=False)
            res["lemma_df"].sort_values(by="Частота", ascending=False).to_excel(writer, sheet_name=f"{label}_Леми", index=False)
            res["top20_df"].to_excel(writer, sheet_name=f"{label}_Топ-20", index=False)
            res["top20_content_df"].to_excel(writer, sheet_name=f"{label}_Топ-20 зміст", index=False)
            # res["latin_df"].to_excel(writer, sheet_name=f"{label}_Латинські", index=False)
            res["abbrev_df"].sort_values(by="Частота", ascending=False).to_excel(writer, sheet_name=f"{label}_Скорочення", index=False)
            res["bigrams_df"].sort_values(by="Частота", ascending=False).to_excel(writer, sheet_name=f"{label}_2-грами", index=False)
            res["trigrams_df"].sort_values(by="Частота", ascending=False).to_excel(writer, sheet_name=f"{label}_3-грами", index=False)

        # Спільні
        make_common_df(common_lemmas, "Лема",
                       result1["lemma_df"].set_index("Лема")["Частота"],
                       result2["lemma_df"].set_index("Лема")["Частота"]
        ).to_excel(writer, sheet_name="Спільні леми", index=False)

        make_common_df(common_bigrams, "2-грам",
                       result1["bigrams_df"].set_index("2-грам")["Частота"],
                       result2["bigrams_df"].set_index("2-грам")["Частота"]
        ).to_excel(writer, sheet_name="Спільні біграми", index=False)

        make_common_df(common_trigrams, "3-грам",
                       result1["trigrams_df"].set_index("3-грам")["Частота"],
                       result2["trigrams_df"].set_index("3-грам")["Частота"]
        ).to_excel(writer, sheet_name="Спільні триграми", index=False)

        # Топ-25: леми
        top_words_1 = result1["main_df"].nlargest(25, "Частота").set_index("Слово")
        top_words_2 = result2["main_df"].nlargest(25, "Частота").set_index("Слово")
        merged_top = top_words_1[["Частота"]].join(top_words_2[["Частота"]], lsuffix=" (ПЦУ)", rsuffix=" (УПЦ)", how="outer").fillna(0)
        merged_top["Відносна (ПЦУ)"] = merged_top["Частота (ПЦУ)"] / merged_top["Частота (ПЦУ)"].sum()
        merged_top["Відносна (УПЦ)"] = merged_top["Частота (УПЦ)"] / merged_top["Частота (УПЦ)"].sum()
        merged_top = merged_top.reset_index()
        merged_top.insert(0, "Номер", range(1, len(merged_top)+1))
        merged_top.to_excel(writer, sheet_name="Топ-25 (порівняння)", index=False)

        # Топ-25: скорочення
        abbrev1 = result1["abbrev_df"].set_index("Скорочення")[["Частота", "Розшифрування"]]
        abbrev2 = result2["abbrev_df"].set_index("Скорочення")[["Частота"]]
        merged_abbrev = abbrev1.join(abbrev2, lsuffix=" (ПЦУ)", rsuffix=" (УПЦ)", how="outer").fillna(0)
        merged_abbrev["Відносна (ПЦУ)"] = merged_abbrev["Частота (ПЦУ)"] / merged_abbrev["Частота (ПЦУ)"].sum()
        merged_abbrev["Відносна (УПЦ)"] = merged_abbrev["Частота (УПЦ)"] / merged_abbrev["Частота (УПЦ)"].sum()
        merged_abbrev = merged_abbrev.reset_index().head(25)
        merged_abbrev.insert(0, "Номер", range(1, len(merged_abbrev)+1))
        merged_abbrev.to_excel(writer, sheet_name="Топ-25 скорочень", index=False)

        # Спільні слова
        words1 = result1["main_df"].set_index("Слово")[["Частота"]]
        words2 = result2["main_df"].set_index("Слово")[["Частота"]]
        common_words = words1.join(words2, lsuffix=" (ПЦУ)", rsuffix=" (УПЦ)", how="inner")
        common_words["Відносна (ПЦУ)"] = common_words["Частота (ПЦУ)"] / common_words["Частота (ПЦУ)"].sum()
        common_words["Відносна (УПЦ)"] = common_words["Частота (УПЦ)"] / common_words["Частота (УПЦ)"].sum()
        common_words = common_words.reset_index()
        common_words.insert(0, "Номер", range(1, len(common_words)+1))
        common_words.to_excel(writer, sheet_name="Спільні слова", index=False)

        autosize_excel_columns(writer, writer.sheets)

# Виклик
start = time.time()
compare_and_save_corpora(pcu_files, upc_files, "ПЦУ", "УПЦ", "27_05.xlsx")
time_now = round(time.time() - start, 2)
print("⏱️ Час виконання:", time_now, "секунд", round(time_now/60, 2), "хвилин")
